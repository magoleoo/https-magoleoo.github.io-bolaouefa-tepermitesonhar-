import json
import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from urllib.request import urlopen
import openpyxl

def slugify(s):
    if not s: return ""
    return str(s).strip().lower().replace(" ", "-")

def normalize(s):
    if s is None:
        return ""
    text = str(s).strip()
    if text == "None":
        return ""
    return text

def format_numeric_cell(value):
    text = normalize(value)
    if text == "":
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return str(number).rstrip("0").rstrip(".")

def to_float(value):
    text = normalize(value)
    if text == "":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0

def normalize_key(value):
    text = normalize(value)
    if text == "":
        return ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower().replace("-", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text

def canonical_team_key(team_name):
    key = normalize_key(team_name)
    aliases = {
        "bayern de munique": "bayern munchen",
        "bayern munchen": "bayern munchen",
        "bayern munich": "bayern munchen",
        "paris saint germain": "paris saint germain",
        "psg": "paris saint germain",
    }
    return aliases.get(key, key)

def split_fixture_label(label):
    clean = normalize(label)
    if clean == "":
        return None, None
    parts = re.split(r"\s+[xX]\s+", clean, maxsplit=1)
    if len(parts) != 2:
        return None, None
    return parts[0].strip(), parts[1].strip()

def is_superclassic_eligible_label(label):
    home, away = split_fixture_label(label)
    if not home or not away:
        return False
    eligible = {
        "real madrid",
        "barcelona",
        "bayern munchen",
        "manchester city",
        "liverpool",
        "chelsea",
        "paris saint germain",
    }
    return canonical_team_key(home) in eligible and canonical_team_key(away) in eligible

def canonical_participant_name(raw_name, participants_by_key):
    name = normalize(raw_name)
    if name == "":
        return ""
    return participants_by_key.get(normalize_key(name), name)

def extract_superclassic_from_sheet(ws, participants_by_key=None):
    participants_by_key = participants_by_key or {}
    fixtures = []
    start_columns = list(range(3, ws.max_column + 1, 6))
    seen_keys = set()

    for row in range(1, ws.max_row + 1):
        for col in start_columns:
            round_label = normalize(ws.cell(row=row, column=col).value)
            if "RODADA" not in round_label.upper():
                continue

            title_row = row + 1
            match_title = normalize(ws.cell(row=title_row, column=col).value)
            if " x " not in match_title.lower():
                continue
            if "xxxx" in normalize_key(match_title):
                # Placeholder da planilha (jogo ainda não definido).
                continue

            key = normalize_key(match_title)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            official_row = None
            for r in range(row + 2, min(ws.max_row + 1, row + 80)):
                if normalize_key(ws.cell(row=r, column=2).value) == "oficial":
                    official_row = r
                    break
            if not official_row:
                continue

            official_h = format_numeric_cell(ws.cell(row=official_row, column=col).value)
            official_sep = normalize_key(ws.cell(row=official_row, column=col + 1).value)
            official_a = format_numeric_cell(ws.cell(row=official_row, column=col + 2).value)
            official_score = (
                f"{official_h}x{official_a}"
                if official_h != "" and official_a != "" and official_sep == "x"
                else ""
            )

            picks = []
            for r in range(row + 2, official_row):
                participant = canonical_participant_name(
                    ws.cell(row=r, column=2).value,
                    participants_by_key,
                )
                if participant == "":
                    continue

                pred_h = format_numeric_cell(ws.cell(row=r, column=col).value)
                pred_sep = normalize_key(ws.cell(row=r, column=col + 1).value)
                pred_a = format_numeric_cell(ws.cell(row=r, column=col + 2).value)
                points = format_numeric_cell(ws.cell(row=r, column=col + 4).value)

                if pred_sep != "x" or pred_h == "" or pred_a == "":
                    continue

                picks.append(
                    {
                        "participant": participant,
                        "pick": f"{pred_h}x{pred_a}",
                        "rank_value": points,
                    }
                )

            fixtures.append(
                {
                    "round_label": round_label,
                    "label": match_title,
                    "official": official_score,
                    "picks": picks,
                }
            )

    return fixtures

def extract_knockout_class_points(
    ws,
    class_header_row,
    class_pick_row_start,
    class_pick_row_end,
    participants_by_key=None,
):
    participants_by_key = participants_by_key or {}
    totals = defaultdict(float)
    for c in range(3, 200, 5):
        header = normalize_key(ws.cell(row=class_header_row, column=c).value)
        if not header.startswith("classificado"):
            continue
        for r in range(class_pick_row_start, class_pick_row_end + 1):
            participant = canonical_participant_name(
                ws.cell(row=r, column=2).value,
                participants_by_key,
            )
            if participant == "" or normalize_key(participant) == "oficial":
                continue
            totals[participant] += to_float(ws.cell(row=r, column=c + 4).value)
    return totals

def extract_phase_summary_points(
    ws,
    start_row,
    end_row,
    name_col,
    points_col,
    participants_by_key=None,
):
    participants_by_key = participants_by_key or {}
    totals = {}
    for row in range(start_row, end_row + 1):
        participant = canonical_participant_name(
            ws.cell(row=row, column=name_col).value,
            participants_by_key,
        )
        if participant == "" or normalize_key(participant) == "oficial":
            continue
        totals[participant] = to_float(ws.cell(row=row, column=points_col).value)
    return totals

def resolve_workbook_path(project_root):
    """
    Fonte oficial do bolão:
    1) Se BOLAO_SOURCE_XLSX_URL estiver configurada, baixa o arquivo temporário.
    2) Caso contrário, usa o arquivo versionado localmente em data/.
    """
    default_path = project_root / "data" / "Bolao_UEFA_25_26_OFICIAL.xlsx"
    source_url = os.getenv("BOLAO_SOURCE_XLSX_URL", "").strip()

    if not source_url:
        return default_path

    download_path = project_root / "data" / "Bolao_UEFA_25_26_OFICIAL.remote.xlsx"
    print(f"Baixando planilha oficial via URL configurada: {source_url}")
    with urlopen(source_url, timeout=60) as response:
        content = response.read()
    if not content:
        raise RuntimeError("Download da planilha retornou arquivo vazio.")
    download_path.write_bytes(content)
    return download_path

def main():
    print("=== INICIANDO PARSER OFICIAL DO EXCEL ===")
    project_root = Path(os.path.dirname(__file__)).parent
    file_path = resolve_workbook_path(project_root)
    if not os.path.exists(file_path):
        print(f"ERRO: Não encontrou {file_path}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 1. PARSE 1a FASE (For Base Participants and Matchday 1)
    ws_1a = wb['1a FASE'] if '1a FASE' in wb.sheetnames else None
    
    participants = []
    part_meta = {} # name -> {artilheiro, garcom, favorito}

    if ws_1a:
        # Participantes estão na Row 2, a partir da Coluna 3.
        # Há uma repetição do bloco de participantes no arquivo oficial, então
        # paramos no primeiro nome repetido para evitar duplicação de palpites.
        r_palpite = 2
        col_start = 3
        seen_participants = set()
        while col_start < 60:
            p_name = normalize(ws_1a.cell(row=r_palpite, column=col_start).value)
            if p_name == "":
                break
            if p_name in seen_participants:
                break
            participants.append({"name": p_name, "col": col_start})
            seen_participants.add(p_name)
            col_start += 1

        # Metadados
        r_art = 5
        r_gar = 6
        r_fav = 7
        for p in participants:
            c = p["col"]
            part_meta[p["name"]] = {
                "artilheiro": normalize(ws_1a.cell(row=r_art, column=c).value),
                "garcom": normalize(ws_1a.cell(row=r_gar, column=c).value),
                "favorito": normalize(ws_1a.cell(row=r_fav, column=c).value)
            }

    participants_by_key = {normalize_key(p["name"]): p["name"] for p in participants}

    # 2. PARSE PHASE TOTALS (sem depender da aba Ranking como fonte primária)
    first_phase_totals = {p["name"]: to_float(ws_1a.cell(row=3, column=p["col"]).value) for p in participants} if ws_1a else {}
    playoff_totals = (
        extract_phase_summary_points(
            wb["PLAYOFF_1aFASE"],
            start_row=98,
            end_row=118,
            name_col=2,
            points_col=6,
            participants_by_key=participants_by_key,
        )
        if "PLAYOFF_1aFASE" in wb.sheetnames
        else {}
    )
    round_of_16_totals = (
        extract_phase_summary_points(
            wb["OITAVAS"],
            start_row=98,
            end_row=118,
            name_col=2,
            points_col=6,
            participants_by_key=participants_by_key,
        )
        if "OITAVAS" in wb.sheetnames
        else {}
    )
    superclassic_totals = (
        extract_phase_summary_points(
            wb["SC"],
            start_row=54,
            end_row=74,
            name_col=2,
            points_col=3,
            participants_by_key=participants_by_key,
        )
        if "SC" in wb.sheetnames
        else {}
    )

    # Metadados e pontos de fases futuras (quartas/semi/final) vindos da aba Ranking.
    ws_rank = wb['Ranking'] if 'Ranking' in wb.sheetnames else None
    ranking_sheet_rows = {}
    if ws_rank:
        for r in range(5, 50):
            raw_name = normalize(ws_rank.cell(row=r, column=4).value)
            if raw_name == "":
                continue
            name = canonical_participant_name(raw_name, participants_by_key)
            ranking_sheet_rows[name] = {
                "sheet_total": to_float(ws_rank.cell(row=r, column=5).value),
                "sheet_first": to_float(ws_rank.cell(row=r, column=6).value),
                "sheet_playoff": to_float(ws_rank.cell(row=r, column=7).value),
                "sheet_round_of_16": to_float(ws_rank.cell(row=r, column=8).value),
                "sheet_quarter": to_float(ws_rank.cell(row=r, column=9).value),
                "sheet_semi": to_float(ws_rank.cell(row=r, column=10).value),
                "sheet_final": to_float(ws_rank.cell(row=r, column=11).value),
                "sheet_superclassic": to_float(ws_rank.cell(row=r, column=12).value),
                "favorite_team": normalize(ws_rank.cell(row=r, column=13).value),
                "assist_pick": normalize(ws_rank.cell(row=r, column=14).value),
                "scorer_pick": normalize(ws_rank.cell(row=r, column=15).value),
            }

    # União de participantes vindos da 1a fase e da aba Ranking.
    participant_names = [p["name"] for p in participants]
    for name in ranking_sheet_rows.keys():
        if name not in participant_names:
            participant_names.append(name)

    ranking_list = []
    deltas = []
    for name in participant_names:
        rank_row = ranking_sheet_rows.get(name, {})
        meta = part_meta.get(name, {})

        first_points = first_phase_totals.get(name, rank_row.get("sheet_first", 0.0))
        playoff_points = playoff_totals.get(name, rank_row.get("sheet_playoff", 0.0))
        round_of_16_points = round_of_16_totals.get(name, rank_row.get("sheet_round_of_16", 0.0))
        superclassic_points = superclassic_totals.get(name, rank_row.get("sheet_superclassic", 0.0))
        quarter_points = rank_row.get("sheet_quarter", 0.0)
        semi_points = rank_row.get("sheet_semi", 0.0)
        final_points = rank_row.get("sheet_final", 0.0)

        total_points = (
            first_points
            + playoff_points
            + round_of_16_points
            + quarter_points
            + semi_points
            + final_points
            + superclassic_points
        )
        total_points = round(total_points, 2)

        sheet_total = rank_row.get("sheet_total")
        delta_vs_sheet = (
            round(total_points - sheet_total, 2)
            if sheet_total is not None
            else 0.0
        )
        if abs(delta_vs_sheet) > 0.01:
            deltas.append((name, delta_vs_sheet))

        favorito = rank_row.get("favorite_team") or meta.get("favorito") or "-"
        artilheiro = rank_row.get("scorer_pick") or meta.get("artilheiro") or "-"
        garcom = rank_row.get("assist_pick") or meta.get("garcom") or "-"

        ranking_list.append({
            "participant_id": slugify(name),
            "name": name,
            "total_points": total_points,
            "first_phase_points": round(first_points, 2),
            "playoff_points": round(playoff_points, 2),
            "round_of_16_points": round(round_of_16_points, 2),
            "quarter_points": round(quarter_points, 2),
            "semi_points": round(semi_points, 2),
            "final_points": round(final_points, 2),
            "superclassic_points": round(superclassic_points, 2),
            "hope_solo_hits": 0,
            "favorite_team": favorito,
            "scorer_pick": artilheiro,
            "assist_pick": garcom,
            "delta_vs_ranking_sheet": delta_vs_sheet,
        })

    if deltas:
        print("[AUDIT][RANKING] Divergências encontradas entre cálculo por fase e aba Ranking:")
        for name, delta in deltas:
            print(f"  - {name}: delta {delta:+.2f}")
    else:
        print("[AUDIT][RANKING] Cálculo por fase bate 100% com a aba Ranking.")

    # Sort Ranking by Total
    ranking_list.sort(key=lambda x: x["total_points"], reverse=True)

    # 3. EXTRAIR PHASES (Palpites)
    phases = {}

    matches_list = []

    def extract_knockout_block(ws, header_row, pick_row_start, pick_row_end, official_row, leg_label):
        fixtures = []
        for c in range(3, 200, 5):
            match_lbl = normalize(ws.cell(row=header_row, column=c).value)
            if not match_lbl or " x " not in match_lbl:
                continue

            official_h = format_numeric_cell(ws.cell(row=official_row, column=c).value)
            official_sep = normalize(ws.cell(row=official_row, column=c + 1).value).lower()
            official_a = format_numeric_cell(ws.cell(row=official_row, column=c + 2).value)
            official_res = (
                f"{official_h}x{official_a}"
                if official_h != "" and official_a != "" and official_sep == "x"
                else "-"
            )

            fixture = {
                "label": match_lbl,
                "leg": leg_label,
                "official": official_res,
                "picks": []
            }

            for r in range(pick_row_start, pick_row_end + 1):
                p_name = canonical_participant_name(
                    ws.cell(row=r, column=2).value,
                    participants_by_key,
                )
                if not p_name or p_name.lower() in ["oficial", "resultado", "resultado oficial", "placar"]:
                    continue

                ph = format_numeric_cell(ws.cell(row=r, column=c).value)
                sep = normalize(ws.cell(row=r, column=c + 1).value).lower()
                pa = format_numeric_cell(ws.cell(row=r, column=c + 2).value)
                pts = format_numeric_cell(ws.cell(row=r, column=c + 4).value)

                if sep == "x" and ph != "" and pa != "":
                    fixture["picks"].append({
                        "participant": p_name,
                        "pick": f"{ph}x{pa}",
                        "rank_value": pts
                    })

            fixtures.append(fixture)
        return fixtures

    def extract_knockout(sheet_name, phase_key):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        fixtures = []
        fixtures.extend(
            extract_knockout_block(
                ws=ws,
                header_row=2,
                pick_row_start=3,
                pick_row_end=23,
                official_row=24,
                leg_label="IDA"
            )
        )
        fixtures.extend(
            extract_knockout_block(
                ws=ws,
                header_row=26,
                pick_row_start=27,
                pick_row_end=47,
                official_row=48,
                leg_label="VOLTA"
            )
        )

        if fixtures:
            phases[phase_key] = {"fixtures": fixtures}

    extract_knockout('PLAYOFF_1aFASE', 'PLAYOFF')
    extract_knockout('OITAVAS', 'ROUND_OF_16')

    # LEAGUE PHASE
    if ws_1a:
        ws_form = wb['Form_PrimeiraFase'] if 'Form_PrimeiraFase' in wb.sheetnames else None
        fixtures_1a = []
        r = 8
        while r < 200:
            md = normalize(ws_1a.cell(row=r, column=1).value)
            if not md: break
            
            off = normalize(ws_1a.cell(row=r, column=2).value)
            
            # Extrair título original do confronto a partir da aba Form_PrimeiraFase
            real_match_title = ""
            if ws_form:
                col_index = r - 4
                real_val = ws_form.cell(row=1, column=col_index).value
                if real_val:
                    real_match_title = str(real_val).strip()
            
            if not real_match_title:
                real_match_title = f"{md} - Jogo {r-7}"

            h = ""
            a = ""
            if " x " in real_match_title:
                h = real_match_title.split(" x ")[0]
                a = real_match_title.split(" x ")[1]
            
            fixture = {
                "matchday": md,
                "label": real_match_title,
                "home": h,
                "away": a,
                "official": off,
                "picks": []
            }
            
            for p in participants:
                pick_val = normalize(ws_1a.cell(row=r, column=p["col"]).value)
                if pick_val:
                    fixture["picks"].append({
                        "participant": p["name"],
                        "pick": pick_val,
                        "rank_value": ""
                    })
            if fixture["picks"]:
                fixtures_1a.append(fixture)
            r += 1
            
        if fixtures_1a:
            phases["LEAGUE"] = {"fixtures": fixtures_1a}

    # 4) Sobrescrever os confrontos superclássicos da 1ª fase com os placares da aba SC.
    sc_fixtures = []
    if "SC" in wb.sheetnames:
        sc_fixtures = extract_superclassic_from_sheet(
            wb["SC"],
            participants_by_key=participants_by_key,
        )

    league_phase = phases.get("LEAGUE", {})
    league_fixtures = league_phase.get("fixtures", [])
    if league_fixtures:
        sc_by_label = {normalize_key(item["label"]): item for item in sc_fixtures}

        expected_superclassic_labels = []
        for fixture in league_fixtures:
            if is_superclassic_eligible_label(fixture.get("label", "")):
                expected_superclassic_labels.append(fixture.get("label", ""))

        expected_superclassic_keys = {normalize_key(label) for label in expected_superclassic_labels}
        missing_superclassic = []
        merged_count = 0

        for fixture in league_fixtures:
            label = fixture.get("label", "")
            if not is_superclassic_eligible_label(label):
                continue
            sc_item = sc_by_label.get(normalize_key(label))
            if not sc_item:
                missing_superclassic.append(label)
                continue

            fixture["picks"] = sc_item.get("picks", [])
            if sc_item.get("official"):
                fixture["official"] = sc_item["official"]
            merged_count += 1

        extra_sc_labels = [
            item["label"]
            for item in sc_fixtures
            if normalize_key(item["label"]) not in expected_superclassic_keys
        ]

        print(
            f"[SC] Superclássicos elegíveis na 1ª fase: {len(expected_superclassic_labels)} | "
            f"com placares importados da aba SC: {merged_count}"
        )
        if missing_superclassic:
            print("[SC][ALERTA] Faltando na aba SC:", ", ".join(missing_superclassic))
        if extra_sc_labels:
            print("[SC][INFO] Confrontos na aba SC fora da grade elegível atual:", ", ".join(extra_sc_labels))


    # Calculate Hope Solo Hits
    hope_solos = {p["name"]: 0 for p in ranking_list}
    for phase_key, phase_data in phases.items():
        for fixture in phase_data["fixtures"]:
            off = fixture.get("official", "")
            if not off or off == "-" or off == "None": continue

            def is_hit(p, o):
                if not p or p == "None": return False
                if p == o: return True
                if "x" in p and "x" in o:
                    try:
                        ph, pa = map(int, p.split("x"))
                        oh, oa = map(int, o.split("x"))
                        if (ph > pa and oh > oa) or (ph < pa and oh < oa) or (ph == pa and oh == oa):
                            return True
                    except: pass
                return False

            correct_picks = []
            for pk in fixture["picks"]:
                if is_hit(pk["pick"], off):
                    correct_picks.append(pk["participant"])
            
            if len(correct_picks) == 1:
                hope_solos[correct_picks[0]] = hope_solos.get(correct_picks[0], 0) + 1
                        
    for r in ranking_list:
        r["hope_solo_hits"] = hope_solos.get(r["name"], 0)

    # OUTPUT JSON
    out_dir = project_root / "api"
    out_dir.mkdir(exist_ok=True)

    ranking_json = {
        "season": 2026,
        "ranking": ranking_list,
        "phases": phases
    }

    with open(out_dir / "ranking.json", "w", encoding="utf-8") as f:
        json.dump(ranking_json, f, ensure_ascii=False, indent=2)

    with open(out_dir / "ranking.js", "w", encoding="utf-8") as f:
        f.write("window.backtestData = " + json.dumps(ranking_json, ensure_ascii=False) + ";\n")

    with open(out_dir / "matches.json", "w", encoding="utf-8") as f:
        json.dump({"matches": matches_list}, f, ensure_ascii=False, indent=2)

    with open(out_dir / "matches.js", "w", encoding="utf-8") as f:
        f.write("window.apiMatchesData = " + json.dumps({"matches": matches_list}, ensure_ascii=False) + ";\n")

    print(f"=== SUCESSO! api/ranking.js E api/matches.js GERAÇÃO CONCLUÍDA ({len(ranking_list)} particpantes) ===")

if __name__ == "__main__":
    main()
